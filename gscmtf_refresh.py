#!/usr/bin/env python3
"""
GSCMTF Dashboard Refresh Script
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Reads live data from GSCMTF_Input_Tracker_v2.xlsx
and injects it into gscmtf_dashboard.html.

HOW TO USE
  1. Open GSCMTF_Input_Tracker_v2.xlsx in Excel and enter your data.
  2. Save the Excel file (Ctrl+S).
  3. Run this script:  python gscmtf_refresh.py
  4. Open gscmtf_dashboard.html in your browser — data is now live.

Both files must be in the same folder as this script.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import json
import re
import sys
from pathlib import Path
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run:  pip install openpyxl")
    sys.exit(1)

# ── File paths (same folder as this script) ──────────────────────────────────
BASE = Path(__file__).parent
EXCEL_FILE = BASE / "GSCMTF_Input_Tracker_v2.xlsx"
HTML_FILE  = BASE / "gscmtf_dashboard.html"

# ── Country metadata (static) ────────────────────────────────────────────────
COUNTRY_META = [
    {"id": "bangladesh", "name": "Bangladesh", "flag": "🇧🇩", "region": "Asia"},
    {"id": "chad",       "name": "Chad",        "flag": "🇹🇩", "region": "West Africa"},
    {"id": "colombia",   "name": "Colombia",    "flag": "🇨🇴", "region": "Latin America"},
    {"id": "drc",        "name": "DRC",         "flag": "🇨🇩", "region": "Central Africa"},
    {"id": "ethiopia",   "name": "Ethiopia",    "flag": "🇪🇹", "region": "East Africa"},
    {"id": "haiti",      "name": "Haiti",       "flag": "🇭🇹", "region": "Caribbean"},
    {"id": "honduras",   "name": "Honduras",    "flag": "🇭🇳", "region": "Central America"},
    {"id": "myanmar",    "name": "Myanmar",     "flag": "🇲🇲", "region": "South-East Asia"},
    {"id": "syria",      "name": "Syria",       "flag": "🇸🇾", "region": "Middle East"},
]

# Excel sheet names must match these
SHEET_NAMES = ["Bangladesh","Chad","Colombia","DRC","Ethiopia",
               "Haiti","Honduras","Myanmar","Syria"]

# Status string → HTML id
STATUS_MAP = {
    "On Track":    "on-track",
    "At Risk":     "at-risk",
    "Delayed":     "delayed",
    "Not Started": "not-started",
    "Completed":   "completed",
}

# ── Row positions in each country sheet (from build_tracker_v2.py) ───────────
# Section 1 — Programmatic (header row 5, data rows 6-22)
R_STATUS        = 6
R_TOTAL_BUDGET  = 7
R_SPENT         = 8   # SUMIF formula — we recompute from detail table
R_COMMITTED     = 9   # SUMIF formula — we recompute from detail table
R_TARGET_BENES  = 12
R_REACHED_BENES = 13
R_TOTAL_MONTHS  = 15
R_COMP_MONTHS   = 16
R_NEXT_DIST     = 18
R_START_DATE    = 19
R_END_DATE      = 20
R_OFFICER       = 21
R_FAO_OFFICE    = 22

# Section 2 — Commodities (header row 24, items rows 25-32)
R_ITEMS_START   = 25
R_ITEMS_END     = 32

# Section 3 — Orders (header row 34, data rows 35-41)
# Rows 35-39 are COUNTIF formulas — we recompute from detail table
R_OFR           = 41   # manual KPI input

# Section 4 — Inventory (header row 43, data rows 44-52)
R_ON_HAND       = 44
R_IN_TRANSIT    = 45
R_CAPACITY      = 47
R_LOC_START     = 49
R_LOC_END       = 52

# Section 5 — Deliveries (header row 54, data rows 55-66)
R_PLANNED       = 55
R_DISPATCHED    = 56
R_RECV_PARTNERS = 57
R_LAST_MILE     = 58
R_TRANSPORT     = 62
R_PARTNER_START = 63
R_PARTNER_END   = 66

# Section 6 — KPIs (header row 68, data rows 69-72)
R_OTD           = 69
R_STOCKOUT      = 70

# Section 7 — Risks (header 74, sub-hdr 75, data rows 76-83)
R_RISK_START    = 76
R_RISK_END      = 83
COL_RISK_LEVEL  = 2   # B
COL_RISK_DESC   = 3   # C
COL_RISK_MITIG  = 4   # D

# Section 8 — Budget Detail (header 89, col-hdr 90, data 91-125)
S8_START        = 91
S8_END          = 125
COL_AMOUNT      = 6   # F
COL_PAY_STATUS  = 8   # H

# Section 9 — Orders Detail (header 127, col-hdr 128, data 129-178)
S9_START        = 129
S9_END          = 178
COL_ORDER_STATUS = 13  # M

# ── Helpers ───────────────────────────────────────────────────────────────────
def cv(ws, row, col=2):
    """Read a cell value, returning None if empty or formula."""
    val = ws.cell(row=row, column=col).value
    if isinstance(val, str) and val.startswith("="):
        return None   # skip formulas — computed separately
    return val

def num(val, default=0):
    """Coerce to number."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default

def intv(val, default=0):
    return int(round(num(val, default)))

def pct_val(val):
    """
    Convert a percentage cell value to an integer 0-100.
    Excel stores decimals (0.75 = 75%) OR integers (75 = 75%).
    """
    v = num(val, 0)
    if 0 < v <= 1:
        return int(round(v * 100))
    return int(round(v))

def fmt_date(val):
    """Format a date/datetime/string for display."""
    if val is None:
        return "TBC"
    if isinstance(val, (datetime, date)):
        return val.strftime("%b %Y")
    return str(val).strip() or "TBC"

def fmt_date_full(val):
    """Format a full date for 'next distribution'."""
    if val is None:
        return "TBC"
    if isinstance(val, (datetime, date)):
        return val.strftime("%d %b %Y")
    return str(val).strip() or "TBC"

def txt(val, default=""):
    if val is None:
        return default
    return str(val).strip()

def collect_list(ws, row_start, row_end, col=2):
    """Collect non-empty text values from a column range."""
    result = []
    for r in range(row_start, row_end + 1):
        v = txt(cv(ws, r, col))
        if v:
            result.append(v)
    return result

# ── SUMIF / COUNTIF in Python ─────────────────────────────────────────────────
def sumif(ws, criteria_col, criteria_val, sum_col, row_start, row_end):
    """Python equivalent of Excel SUMIF."""
    total = 0.0
    for r in range(row_start, row_end + 1):
        crit = txt(ws.cell(row=r, column=criteria_col).value)
        if crit.lower() == criteria_val.lower():
            total += num(ws.cell(row=r, column=sum_col).value, 0)
    return total

def countif(ws, col, criteria_val, row_start, row_end):
    """Python equivalent of Excel COUNTIF."""
    count = 0
    for r in range(row_start, row_end + 1):
        v = txt(ws.cell(row=r, column=col).value)
        if v.lower() == criteria_val.lower():
            count += 1
    return count

# ── Read one country sheet ────────────────────────────────────────────────────
def read_country(ws, meta):
    # ── Section 1: Programmatic ─────────────────────────────────────────────
    status_raw   = txt(cv(ws, R_STATUS), "Not Started")
    status       = STATUS_MAP.get(status_raw, "not-started")
    total_budget = num(cv(ws, R_TOTAL_BUDGET), 0)
    target_benes = intv(cv(ws, R_TARGET_BENES))
    reached_benes= intv(cv(ws, R_REACHED_BENES))
    total_months = intv(cv(ws, R_TOTAL_MONTHS))
    comp_months  = intv(cv(ws, R_COMP_MONTHS))
    next_dist    = fmt_date_full(cv(ws, R_NEXT_DIST))
    start_date   = fmt_date(cv(ws, R_START_DATE))
    end_date     = fmt_date(cv(ws, R_END_DATE))
    officer      = txt(cv(ws, R_OFFICER))
    fao_office   = txt(cv(ws, R_FAO_OFFICE))

    # Budget: compute SUMIF from Section 8
    spent     = sumif(ws, COL_PAY_STATUS, "Spent",     COL_AMOUNT, S8_START, S8_END)
    committed = sumif(ws, COL_PAY_STATUS, "Committed", COL_AMOUNT, S8_START, S8_END)

    # ── Section 2: Commodities ───────────────────────────────────────────────
    items = collect_list(ws, R_ITEMS_START, R_ITEMS_END)
    if not items:
        items = ["—"]

    # ── Section 3: Orders — COUNTIF from Section 9 ──────────────────────────
    o_pending    = countif(ws, COL_ORDER_STATUS, "Pending",       S9_START, S9_END)
    o_confirmed  = countif(ws, COL_ORDER_STATUS, "Confirmed",     S9_START, S9_END)
    o_inprod     = countif(ws, COL_ORDER_STATUS, "In Production", S9_START, S9_END)
    o_shipped    = countif(ws, COL_ORDER_STATUS, "Shipped",       S9_START, S9_END)
    o_delivered  = countif(ws, COL_ORDER_STATUS, "Delivered",     S9_START, S9_END)
    ofr          = pct_val(cv(ws, R_OFR))

    # ── Section 4: Inventory ─────────────────────────────────────────────────
    on_hand   = intv(cv(ws, R_ON_HAND))
    in_transit= intv(cv(ws, R_IN_TRANSIT))
    capacity  = intv(cv(ws, R_CAPACITY)) or 1000
    locations = collect_list(ws, R_LOC_START, R_LOC_END)
    if not locations:
        locations = ["—"]

    # ── Section 5: Deliveries ─────────────────────────────────────────────────
    planned     = intv(cv(ws, R_PLANNED))
    dispatched  = intv(cv(ws, R_DISPATCHED))
    recv_part   = intv(cv(ws, R_RECV_PARTNERS))
    last_mile   = intv(cv(ws, R_LAST_MILE))
    transport   = txt(cv(ws, R_TRANSPORT), "—")
    partners    = collect_list(ws, R_PARTNER_START, R_PARTNER_END)
    if not partners:
        partners = ["—"]

    # ── Section 6: KPIs ──────────────────────────────────────────────────────
    otd          = pct_val(cv(ws, R_OTD))
    stockout     = txt(cv(ws, R_STOCKOUT), "—")

    # ── Section 7: Risks ─────────────────────────────────────────────────────
    risks = []
    for r in range(R_RISK_START, R_RISK_END + 1):
        level  = txt(ws.cell(row=r, column=COL_RISK_LEVEL).value).lower()
        desc   = txt(ws.cell(row=r, column=COL_RISK_DESC).value)
        mitig  = txt(ws.cell(row=r, column=COL_RISK_MITIG).value)
        if desc:
            risks.append({
                "level":      level or "low",
                "desc":       desc,
                "mitigation": mitig,
                "owner":      ""
            })

    # ── Assemble country object ───────────────────────────────────────────────
    return {
        "id":     meta["id"],
        "name":   meta["name"],
        "flag":   meta["flag"],
        "region": meta["region"],
        "status": status,
        "programmatic": {
            "budget": {
                "total":     total_budget,
                "spent":     spent,
                "committed": committed,
            },
            "beneficiaries": {
                "target":  target_benes,
                "reached": reached_benes,
            },
            "distributions": {
                "totalMonths":     total_months,
                "completedMonths": comp_months,
                "nextDist":        next_dist,
            },
            "items":       items,
            "startDate":   start_date,
            "endDate":     end_date,
            "projectOfficer": officer,
            "faoOffice":      fao_office,
        },
        "sc": {
            "orders": {
                "pending":      o_pending,
                "confirmed":    o_confirmed,
                "inProduction": o_inprod,
                "shipped":      o_shipped,
                "delivered":    o_delivered,
            },
            "inventory": {
                "onHand":    on_hand,
                "inTransit": in_transit,
                "capacity":  capacity,
                "unit":      "MT",
                "locations": locations,
            },
            "deliveries": {
                "planned":           planned,
                "dispatched":        dispatched,
                "receivedByPartners":recv_part,
                "lastMile":          last_mile,
            },
            "partners":  partners,
            "transport": transport,
            "kpis": {
                "orderFulfillment": ofr,
                "onTimeDelivery":   otd,
                "stockoutRisk":     stockout or "—",
            },
            "risks": risks,
        },
    }

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("━" * 60)
    print("  GSCMTF Dashboard Refresh")
    print("━" * 60)

    # Validate files exist
    if not EXCEL_FILE.exists():
        print(f"\n❌  Excel file not found:\n    {EXCEL_FILE}")
        print("    Make sure both files are in the same folder.")
        sys.exit(1)
    if not HTML_FILE.exists():
        print(f"\n❌  HTML file not found:\n    {HTML_FILE}")
        sys.exit(1)

    # Load workbook (without data_only so we read raw input cells;
    # SUMIF/COUNTIF are computed in Python from the detail tables)
    print(f"\n📂  Reading Excel: {EXCEL_FILE.name}")
    wb = load_workbook(EXCEL_FILE, data_only=False)

    countries_data = []
    for meta, sheet_name in zip(COUNTRY_META, SHEET_NAMES):
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️   Sheet '{sheet_name}' not found — skipping")
            continue
        ws = wb[sheet_name]
        country_obj = read_country(ws, meta)
        countries_data.append(country_obj)

        # Summary line per country
        p = country_obj["programmatic"]
        sc = country_obj["sc"]
        orders_total = sum(sc["orders"].values())
        print(f"  ✅  {meta['name']:<14}  "
              f"Status: {country_obj['status']:<12}  "
              f"Budget: ${p['budget']['total']:>10,.0f}  "
              f"Orders: {orders_total}")

    # Serialise to JS
    js_data = json.dumps(countries_data, indent=2, ensure_ascii=False)
    new_block = (
        "// ==GSCMTF_DATA_START==\n"
        "const countries = \n"
        f"{js_data};\n"
        "// ==GSCMTF_DATA_END=="
    )

    # Inject into HTML
    print(f"\n💉  Injecting data into: {HTML_FILE.name}")
    html = HTML_FILE.read_text(encoding="utf-8")
    pattern = r"// ==GSCMTF_DATA_START==.*?// ==GSCMTF_DATA_END=="
    if not re.search(pattern, html, re.DOTALL):
        print("❌  Data markers not found in HTML file.")
        print("    Expected: // ==GSCMTF_DATA_START== and // ==GSCMTF_DATA_END==")
        sys.exit(1)

    new_html = re.sub(pattern, new_block, html, flags=re.DOTALL)

    # Add last-refresh timestamp to header
    ts = datetime.now().strftime("%d %b %Y %H:%M")
    new_html = re.sub(
        r'Last updated: <b>.*?</b>',
        f'Last updated: <b>{ts}</b>',
        new_html
    )

    HTML_FILE.write_text(new_html, encoding="utf-8")

    # Summary
    total_budget  = sum(c["programmatic"]["budget"]["total"]  for c in countries_data)
    total_benes   = sum(c["programmatic"]["beneficiaries"]["target"] for c in countries_data)
    total_orders  = sum(sum(c["sc"]["orders"].values()) for c in countries_data)
    on_track = sum(1 for c in countries_data if c["status"] == "on-track")
    at_risk  = sum(1 for c in countries_data if c["status"] == "at-risk")
    delayed  = sum(1 for c in countries_data if c["status"] == "delayed")

    print(f"\n{'━'*60}")
    print(f"  ✅  Dashboard refreshed successfully!")
    print(f"{'━'*60}")
    print(f"  Countries processed : {len(countries_data)}")
    print(f"  Total Budget        : ${total_budget:,.0f}")
    print(f"  Total Beneficiaries : {total_benes:,}")
    print(f"  Total Orders        : {total_orders}")
    print(f"  Status breakdown    : ✅ On Track: {on_track}  "
          f"⚠️ At Risk: {at_risk}  🔴 Delayed: {delayed}")
    print(f"\n  Open {HTML_FILE.name} in your browser to view the dashboard.")
    print(f"{'━'*60}\n")


if __name__ == "__main__":
    main()
